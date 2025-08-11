import os
import pandas as pd
from openpyxl import load_workbook


# for any input file
    # find matching input row
    # find rent column
    # loop thru rent column and copy rents

print('==============================================================')
print('Welcome! Make sure to place your files in the "data" directory')
print('! NOTE - IF FIRST TIME RUN, MAKE COPIES OF ALL FILES BEFOREHAND.')
print('-----------------------------------')
print('Please select an option:')
print('1. Update rents')
print('2. Add new property')
option = input('Your option: ')

if option == '1':
    print('You chose to update rent values.')
    print('-----------------------------------')

    wb = load_workbook('Lbk Comp.xlsx')
    ws = wb['Comps']

    # Will run once for each file in data directory
    for file in os.listdir('data'):
        print('PROCESSING FILE: ', file)

        # Open "Lbk Comp" Excel file with the pandas library
        df = pd.read_excel('data/' + file)
        # Include only columns we care about
        df = df[['Avg SF', 'Avg Asking Rent/Unit']]

        # Exclude all data below blank row
        null_row = df.index[df.isnull().all(axis=1)][0]
        df = df.loc[:null_row - 1]

        # Now, we need to find the matching block in "Lbk Comp"
        # We do this by cycling thru col G to find matching SQFT
        for cell in ws['D']:
            if cell.value == int(df['Avg SF'].iloc[0]):
                # We find the first row of the block
                target = cell.row

        # First, we create a column that starts at our target cell -
        # The cell in Comp that matches SQFT in PID file
        col_C = ws['C'][target - 1:]
        # Then, we go thru each row in the apartment block and
        # update the rent using the data from the PID file
        for i, cell in enumerate(col_C):
            if i < len(df):
                cell.value = df['Avg Asking Rent/Unit'].iloc[i]


    # Save changes to file
    wb.save('Lbk Comp.xlsx')

if option == '2':
    print('This option is not implemented yet.')


